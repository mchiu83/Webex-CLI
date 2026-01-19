# Copyright (c) 2026 Ming Chiu
# Licensed under the MIT License - see LICENSE file for details

import os
import glob
import openpyxl
import xlrd
import re
from libraries.add_device import PHONE_MODELS, COLLAB_MODELS

def find_aso_import_file():
    """Find Excel file with prefix 'aso_import' in bulk directory"""
    bulk_dir = 'bulk'
    
    if not os.path.exists(bulk_dir):
        return None
    
    patterns = [os.path.join(bulk_dir, 'aso_import*.xlsx'), 
                os.path.join(bulk_dir, 'aso_import*.xls')]
    
    for pattern in patterns:
        files = glob.glob(pattern)
        if files:
            return files[0]
    
    return None

def read_excel_sheet(filepath, sheet_name):
    """Read data from specific Excel sheet"""
    try:
        if filepath.endswith('.xlsx'):
            import warnings
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            wb.close()
            return data
        elif filepath.endswith('.xls'):
            wb = xlrd.open_workbook(filepath, formatting_info=False)
            ws = wb.sheet_by_name(sheet_name)
            data = []
            for row_idx in range(ws.nrows):
                data.append(ws.row_values(row_idx))
            return data
    except Exception as e:
        print(f"Error reading sheet '{sheet_name}': {str(e)}")
        return None

def validate_excel_file(filepath):
    """Validate Excel file structure and required tabs"""
    print(f"\nValidating Excel file: {filepath}")
    print(f"{'='*60}")
    
    required_tabs = ['Webex Users', 'Webex Side Cars', 'Webex Auto Attendant', 'Webex Hunt Groups']
    validation_results = []
    additional_tabs = []
    
    try:
        # Determine file type and load workbook
        if filepath.endswith('.xlsx'):
            wb = openpyxl.load_workbook(filepath, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        elif filepath.endswith('.xls'):
            wb = xlrd.open_workbook(filepath)
            sheet_names = wb.sheet_names()
        else:
            print(f"Status: FAILED - Unsupported file format")
            return False, None
        
        # Validation 1: Check for required tabs
        print(f"\nValidation 1: Checking required tabs...")
        all_required_present = True
        
        for tab in required_tabs:
            if tab in sheet_names:
                print(f"  [{tab}] - Status: PASS")
                validation_results.append({'tab': tab, 'status': 'PASS'})
            else:
                print(f"  [{tab}] - Status: FAILED (Missing)")
                validation_results.append({'tab': tab, 'status': 'FAILED'})
                all_required_present = False
        
        if not all_required_present:
            print(f"\nOverall Status: FAILED - Missing required tabs")
            return False, None
        
        # Validation 2: Check for additional tabs
        print(f"\nValidation 2: Checking for additional tabs...")
        additional_tabs = [tab for tab in sheet_names if tab not in required_tabs]
        
        if not additional_tabs:
            print(f"  Status: FAILED - No additional tabs found (at least one required)")
            return False, None
        else:
            print(f"  Status: PASS - Found {len(additional_tabs)} additional tab(s)")
            for tab in additional_tabs:
                print(f"    - {tab}")
        
        print(f"\n{'='*60}")
        print(f"Overall Validation Status: PASS")
        print(f"{'='*60}")
        
        return True, additional_tabs
        
    except Exception as e:
        print(f"\nStatus: FAILED - Error reading Excel file: {str(e)}")
        return False, None

def validate_location(api, filepath):
    """Validation 3: Infer and validate location from Webex Users sheet"""
    print(f"\nValidation 3: Inferring and validating location...")
    
    # Read Webex Users sheet
    users_data = read_excel_sheet(filepath, 'Webex Users')
    if not users_data:
        print(f"  Status: FAILED - Could not read 'Webex Users' sheet")
        return None
    
    # Find Location Name column
    headers = users_data[0]
    location_col_idx = None
    
    for idx, header in enumerate(headers):
        if header and 'Location Name' in str(header):
            location_col_idx = idx
            break
    
    if location_col_idx is None:
        print(f"  Status: FAILED - 'Location Name' column not found in 'Webex Users' sheet")
        return None
    
    # Get location name from first data row
    if len(users_data) < 2:
        print(f"  Status: FAILED - No data rows found in 'Webex Users' sheet")
        return None
    
    inferred_location = None
    for row in users_data[1:]:
        if row[location_col_idx]:
            inferred_location = str(row[location_col_idx]).strip()
            break
    
    if not inferred_location:
        print(f"  Status: FAILED - No location name found in data rows")
        return None
    
    print(f"  Inferred Location: {inferred_location}")
    
    # Fetch telephony locations from Webex API
    print(f"  Fetching telephony locations from Webex API...")
    locations_result = api.call("GET", "telephony/config/locations", params={"orgId": api.org_id})
    
    if "error" in locations_result:
        print(f"  Status: FAILED - Error fetching locations: {locations_result['error']}")
        return None
    
    locations = locations_result.get("locations", [])
    
    # Find matching location (case-insensitive)
    matched_location = None
    for loc in locations:
        if loc.get('name', '').lower() == inferred_location.lower():
            matched_location = loc
            break
    
    if matched_location:
        print(f"  Status: PASS - Location match found")
        print(f"  Location ID: {matched_location['id']}")
        print(f"  Location Name: {matched_location['name']}")
        
        # Get calling line ID phone number
        calling_line_id = matched_location.get('callingLineId', {})
        phone_number = calling_line_id.get('phoneNumber')
        if phone_number:
            print(f"  Calling Line ID: {phone_number}")
        else:
            print(f"  Calling Line ID: None")
        
        # Check location outgoing permissions
        print(f"\n  Checking location outgoing permissions...")
        location_id = matched_location['id']
        perm_result = api.call("GET", f"telephony/config/locations/{location_id}/outgoingPermission", 
                              params={"orgId": api.org_id})
        
        if "error" in perm_result:
            print(f"  Status: FAILED - Error fetching outgoing permissions: {perm_result['error']}")
            return None
        
        permissions = perm_result.get('callingPermissions', [])
        
        # Display permissions table
        print(f"\n  Location Outgoing Permissions:")
        print(f"  {'Call Type':<35} {'Action':<10} {'Transfer'}")
        print(f"  {'-'*55}")
        
        mismatches = []
        for perm in permissions:
            call_type = perm.get('callType', '')
            action = perm.get('action', '')
            transfer = perm.get('transferEnabled', False)
            print(f"  {call_type:<35} {action:<10} {transfer}")
            
            # Skip ignored call types
            if call_type in ['CASUAL', 'URL_DIALING', 'UNKNOWN']:
                continue
            
            # Check expected values
            if call_type == 'INTERNAL_CALL':
                if action != 'ALLOW' or transfer != True:
                    mismatches.append(call_type)
            else:
                if action != 'BLOCK' or transfer != False:
                    mismatches.append(call_type)
        
        if not mismatches:
            print(f"\n  Status: PASS - Location permissions match expected configuration")
        else:
            print(f"\n  Status: WARNING - Location permissions do not match expected configuration")
            print(f"  Mismatched call types: {', '.join(mismatches)}")
            
            modify = input("\n  Modify location outgoing permissions to default? (y/n): ").strip().lower()
            if modify == 'y':
                print(f"  Updating location outgoing permissions...")
                
                update_data = {
                    "callingPermissions": [
                        {"callType": "INTERNAL_CALL", "action": "ALLOW", "transferEnabled": True},
                        {"callType": "TOLL_FREE", "action": "BLOCK", "transferEnabled": False},
                        {"callType": "INTERNATIONAL", "action": "BLOCK", "transferEnabled": False},
                        {"callType": "OPERATOR_ASSISTED", "action": "BLOCK", "transferEnabled": False},
                        {"callType": "CHARGEABLE_DIRECTORY_ASSISTED", "action": "BLOCK", "transferEnabled": False},
                        {"callType": "SPECIAL_SERVICES_I", "action": "BLOCK", "transferEnabled": False},
                        {"callType": "SPECIAL_SERVICES_II", "action": "BLOCK", "transferEnabled": False},
                        {"callType": "PREMIUM_SERVICES_I", "action": "BLOCK", "transferEnabled": False},
                        {"callType": "PREMIUM_SERVICES_II", "action": "BLOCK", "transferEnabled": False},
                        {"callType": "NATIONAL", "action": "BLOCK", "transferEnabled": False}
                    ]
                }
                
                update_result = api.call("PUT", f"telephony/config/locations/{location_id}/outgoingPermission",
                                        data=update_data, params={"orgId": api.org_id})
                
                if "error" in update_result:
                    print(f"  Status: FAILED - Error updating permissions: {update_result['error']}")
                    return None
                else:
                    print(f"  Status: SUCCESS - Location outgoing permissions updated")
            else:
                print(f"  Proceeding without modifying location permissions...")
        
        # Store location data
        location_data = {
            'id': matched_location['id'],
            'name': matched_location['name'],
            'callingLineId': phone_number
        }
        
        return location_data
    else:
        print(f"  Status: FAILED - Location '{inferred_location}' not found in organization.")
        print(f"  Be sure the location is created before proceeding.")
        return None

def validate_webex_users_data(filepath):
    """Validation 4: Validate Webex Users sheet data"""
    print(f"\nValidation 4: Validating Webex Users data...")
    
    # Read Webex Users sheet
    users_data = read_excel_sheet(filepath, 'Webex Users')
    if not users_data or len(users_data) < 2:
        print(f"  Status: FAILED - Could not read 'Webex Users' sheet or no data rows")
        return False
    
    headers = users_data[0]
    data_rows = users_data[1:]
    
    # Define optional columns (A=0, B=1, D=3, F=5, G=6, I=8, N=13, O=14, P=15, Q=16, R=17, S=18)
    optional_cols = [0, 1, 3, 5, 6, 8, 13, 14, 15, 16, 17, 18]
    
    # Track MAC addresses for duplicate check
    mac_addresses = []
    
    # Validate each row
    for row_idx, row in enumerate(data_rows, start=2):
        # Step 1: Check mandatory columns (only A through S, indices 0-18)
        for col_idx in range(min(19, len(headers))):
            if col_idx not in optional_cols:
                if col_idx >= len(row) or not row[col_idx] or str(row[col_idx]).strip() == '':
                    col_name = str(headers[col_idx]).replace('\n', ' ').replace('\r', ' ') if col_idx < len(headers) else f"Column {chr(65 + col_idx)}"
                    print(f"  Status: FAILED - Row {row_idx}: Missing required value in '{col_name}'")
                    print(f"  Row data: {row[:19]}")  # Only show columns A-S
                    return False
        
        # Step 2: Validate MAC address in column L (index 11)
        if len(row) <= 11 or not row[11] or str(row[11]).strip() == '':
            col_name = str(headers[11]).replace('\n', ' ').replace('\r', ' ') if 11 < len(headers) else "Column L"
            print(f"  Status: FAILED - Row {row_idx}: Missing MAC address in '{col_name}'")
            print(f"  Row data: {row}")
            return False
        
        mac_raw = str(row[11]).strip()
        # Remove dashes, spaces, colons and validate format
        mac_clean = re.sub(r'[-:\s]', '', mac_raw).upper()
        if not re.match(r'^[0-9A-F]{12}$', mac_clean):
            col_name = str(headers[11]).replace('\n', ' ').replace('\r', ' ') if 11 < len(headers) else "Column L"
            print(f"  Status: FAILED - Row {row_idx}: Invalid MAC address format in '{col_name}': '{mac_raw}'")
            print(f"  Row data: {row}")
            return False
        
        mac_addresses.append(mac_clean)
        
        # Step 3: Validate column J (index 9) - must be 'non-user' or 'user'
        if len(row) <= 9 or not row[9]:
            col_name = str(headers[9]).replace('\n', ' ').replace('\r', ' ') if 9 < len(headers) else "Column J"
            print(f"  Status: FAILED - Row {row_idx}: Missing value in '{col_name}'")
            print(f"  Row data: {row}")
            return False
        
        col_j_value = str(row[9]).strip().lower()
        if col_j_value not in ['non-user', 'user']:
            col_name = str(headers[9]).replace('\n', ' ').replace('\r', ' ') if 9 < len(headers) else "Column J"
            print(f"  Status: FAILED - Row {row_idx}: '{col_name}' must be 'non-user' or 'user', got '{row[9]}'")
            print(f"  Row data: {row}")
            return False
        
        # Step 4: Validate column E (index 4) - must be numeric
        if len(row) > 4 and row[4] and str(row[4]).strip() != '':
            try:
                float(row[4])
            except ValueError:
                col_name = str(headers[4]).replace('\n', ' ').replace('\r', ' ') if 4 < len(headers) else "Column E"
                print(f"  Status: FAILED - Row {row_idx}: '{col_name}' must be numeric, got '{row[4]}'")
                print(f"  Row data: {row}")
                return False
        
        # Step 5: Validate column D (index 3) - must be empty or 10 digit numeric
        if len(row) > 3 and row[3] and str(row[3]).strip() != '':
            col_d_value = str(row[3]).strip()
            if not col_d_value.isdigit() or len(col_d_value) != 10:
                col_name = str(headers[3]).replace('\n', ' ').replace('\r', ' ') if 3 < len(headers) else "Column D"
                print(f"  Status: FAILED - Row {row_idx}: '{col_name}' must be empty or 10 digit numeric, got '{row[3]}'")
                print(f"  Row data: {row}")
                return False
        
        # Step 6: Validate column O (index 14) - must be empty or numeric <= 15
        if len(row) > 14 and row[14] and str(row[14]).strip() != '':
            try:
                col_o_value = float(row[14])
                if col_o_value > 15:
                    col_name = str(headers[14]).replace('\n', ' ').replace('\r', ' ') if 14 < len(headers) else "Column O"
                    print(f"  Status: FAILED - Row {row_idx}: '{col_name}' must be <= 15, got '{row[14]}'")
                    print(f"  Row data: {row}")
                    return False
            except ValueError:
                col_name = str(headers[14]).replace('\n', ' ').replace('\r', ' ') if 14 < len(headers) else "Column O"
                print(f"  Status: FAILED - Row {row_idx}: '{col_name}' must be numeric, got '{row[14]}'")
                print(f"  Row data: {row}")
                return False
        
        # Step 7: Validate column P (index 15) and R (index 17) - must be 'yes', 'no', or empty
        for col_idx, col_letter in [(15, 'P'), (17, 'R')]:
            if len(row) > col_idx and row[col_idx] and str(row[col_idx]).strip() != '':
                col_value = str(row[col_idx]).strip().lower()
                if col_value not in ['yes', 'no']:
                    col_name = str(headers[col_idx]).replace('\n', ' ').replace('\r', ' ') if col_idx < len(headers) else f"Column {col_letter}"
                    print(f"  Status: FAILED - Row {row_idx}: '{col_name}' must be 'yes', 'no', or empty, got '{row[col_idx]}'")
                    print(f"  Row data: {row}")
                    return False
        
        # Step 8: Validate column N (index 13) and Q (index 16) - must be empty or numeric
        for col_idx, col_letter in [(13, 'N'), (16, 'Q')]:
            if len(row) > col_idx and row[col_idx] and str(row[col_idx]).strip() != '':
                try:
                    float(row[col_idx])
                except ValueError:
                    col_name = str(headers[col_idx]).replace('\n', ' ').replace('\r', ' ') if col_idx < len(headers) else f"Column {col_letter}"
                    print(f"  Status: FAILED - Row {row_idx}: '{col_name}' must be numeric, got '{row[col_idx]}'")
                    print(f"  Row data: {row}")
                    return False
    
    # Check for duplicate MAC addresses
    if len(mac_addresses) != len(set(mac_addresses)):
        duplicates = [mac for mac in set(mac_addresses) if mac_addresses.count(mac) > 1]
        print(f"  Status: FAILED - Duplicate MAC addresses found: {', '.join(duplicates)}")
        return False
    
    print(f"  Status: PASS - All {len(data_rows)} rows validated successfully")
    return True

def validate_available_numbers(api, location_data, filepath):
    """Validation 5: Validate phone numbers in Column D against available location numbers"""
    print(f"\nValidation 5: Validating phone number availability...")
    
    # Get available numbers for location
    print(f"  Fetching available numbers for location...")
    location_id = location_data['id']
    numbers_result = api.call("GET", f"telephony/config/locations/{location_id}/availableNumbers", 
                             params={"orgId": api.org_id})
    
    if "error" in numbers_result:
        print(f"  Status: FAILED - Error fetching available numbers: {numbers_result['error']}")
        return False
    
    phone_numbers = numbers_result.get("phoneNumbers", [])
    
    # Filter for available numbers (no owner, not main number, state is ACTIVE)
    available_location_numbers = []
    for num in phone_numbers:
        if (not num.get('owner') and 
            not num.get('isMainNumber', False) and 
            num.get('state') == 'ACTIVE'):
            available_location_numbers.append(num['phoneNumber'])
    
    print(f"  Found {len(available_location_numbers)} available numbers")
    
    # Read Webex Users sheet
    users_data = read_excel_sheet(filepath, 'Webex Users')
    if not users_data or len(users_data) < 2:
        print(f"  Status: FAILED - Could not read 'Webex Users' sheet")
        return False
    
    headers = users_data[0]
    data_rows = users_data[1:]
    
    # Check column D (index 3) phone numbers
    for row_idx, row in enumerate(data_rows, start=2):
        if len(row) > 3 and row[3] and str(row[3]).strip() != '':
            # Get 10-digit number from column D
            phone_10digit = str(row[3]).strip()
            
            # Convert to E.164 format (+1XXXXXXXXXX)
            phone_e164 = f"+1{phone_10digit}"
            
            # Check if number is in available list
            if phone_e164 not in available_location_numbers:
                col_name = str(headers[3]).replace('\n', ' ').replace('\r', ' ') if 3 < len(headers) else "Column D"
                print(f"  Status: FAILED - Row {row_idx}: Phone number '{phone_10digit}' in '{col_name}' is not available")
                print(f"  The number {phone_e164} is not found in available PSTN numbers for this location.")
                print(f"  Please check that the assigned number is available as a PSTN number in the Webex Location.")
                return False
    
    print(f"  Status: PASS - All phone numbers are available")
    return True

def create_workspace_from_row(api, location_data, row, headers):
    """Create workspace from Excel row data"""
    # Extract data from row
    display_name = str(row[12]).strip() if len(row) > 12 else ""  # Column M
    phone_number = str(row[3]).strip() if len(row) > 3 and row[3] else None  # Column D
    extension = str(row[4]).strip() if len(row) > 4 else ""  # Column E
    device_model = str(row[10]).strip() if len(row) > 10 else ""  # Column K
    mac_address = str(row[11]).strip() if len(row) > 11 else ""  # Column L
    
    # Determine supported devices based on model
    if device_model in PHONE_MODELS:
        supported_devices = "phones"
    elif device_model in COLLAB_MODELS:
        supported_devices = "collaborationDevices"
    else:
        supported_devices = "collaborationDevices"  # Default
    
    # Build workspace data
    data = {
        "displayName": display_name,
        "orgId": api.org_id,
        "type": "notSet",
        "supportedDevices": supported_devices,
        "locationId": location_data['id'],
        "calling": {
            "type": "webexCalling",
            "webexCalling": {
                "extension": extension,
                "locationId": location_data['id']
            }
        }
    }
    
    # Add phone number if provided (convert to E.164)
    if phone_number and phone_number.isdigit() and len(phone_number) == 10:
        data["calling"]["webexCalling"]["phoneNumber"] = f"+1{phone_number}"
    
    # Create workspace
    result = api.call("POST", "workspaces", data=data)
    
    if "error" in result:
        return None, result['error']
    
    workspace_id = result.get("id")
    
    # Add device via MAC address
    if workspace_id and mac_address and device_model:
        # Clean MAC address
        mac_clean = re.sub(r'[-:\s]', '', mac_address).upper()
        mac_formatted = ':'.join(mac_clean[i:i+2] for i in range(0, 12, 2))
        
        device_data = {
            "mac": mac_formatted,
            "model": device_model,
            "workspaceId": workspace_id
        }
        
        device_result = api.call("POST", "devices", data=device_data, params={"orgId": api.org_id})
        
        if "error" in device_result:
            return workspace_id, f"Workspace created but device failed: {device_result['error']}"
    
    return workspace_id, None

def configure_call_forwarding(api, workspace_id, row):
    """Configure call forwarding and business continuity for workspace"""
    forward_no_answer = str(row[13]).strip() if len(row) > 13 and row[13] else None  # Column N
    num_rings = str(row[14]).strip() if len(row) > 14 and row[14] else "3"  # Column O
    forward_disconnect = str(row[16]).strip() if len(row) > 16 and row[16] else None  # Column Q
    
    # Check if any forwarding is needed
    if not forward_no_answer and not forward_disconnect:
        return None
    
    # Always include callForwarding structure
    data = {
        "callForwarding": {
            "always": {},
            "busy": {"enabled": False},
            "noAnswer": {
                "enabled": True if forward_no_answer else False,
                "destination": forward_no_answer if forward_no_answer else "",
                "numberOfRings": int(float(num_rings)) if forward_no_answer else 3
            }
        }
    }
    
    # Add businessContinuity if Column Q has value
    if forward_disconnect:
        data["businessContinuity"] = {
            "enabled": True,
            "destination": forward_disconnect
        }
    
    # Send PUT request
    result = api.call("PUT", f"workspaces/{workspace_id}/features/callForwarding", 
                     data=data, params={"orgId": api.org_id})
    
    if "error" in result:
        return result['error']
    
    return None

def configure_outgoing_permission(api, workspace_id, row):
    """Configure outgoing calling permissions for workspace"""
    calling_permission = str(row[18]).strip().lower() if len(row) > 18 and row[18] else None  # Column S
    
    # Only configure if Column S is 'custom'
    if calling_permission != 'custom':
        return None, False
    
    data = {
        "useCustomEnabled": True,
        "useCustomPermissions": True,
        "callingPermissions": [
            {"callType": "INTERNAL_CALL", "action": "ALLOW", "transferEnabled": True},
            {"callType": "TOLL_FREE", "action": "ALLOW", "transferEnabled": True},
            {"callType": "NATIONAL", "action": "ALLOW", "transferEnabled": True},
            {"callType": "INTERNATIONAL", "action": "BLOCK", "transferEnabled": False},
            {"callType": "OPERATOR_ASSISTED", "action": "BLOCK", "transferEnabled": False},
            {"callType": "CHARGEABLE_DIRECTORY_ASSISTED", "action": "BLOCK", "transferEnabled": False},
            {"callType": "SPECIAL_SERVICES_I", "action": "BLOCK", "transferEnabled": False},
            {"callType": "SPECIAL_SERVICES_II", "action": "BLOCK", "transferEnabled": False},
            {"callType": "PREMIUM_SERVICES_I", "action": "BLOCK", "transferEnabled": False},
            {"callType": "PREMIUM_SERVICES_II", "action": "BLOCK", "transferEnabled": False}
        ]
    }
    
    result = api.call("PUT", f"workspaces/{workspace_id}/features/outgoingPermission", 
                     data=data, params={"orgId": api.org_id})
    
    if "error" in result:
        return result['error'], True
    
    return None, True

def configure_side_car_speed_dials(api, workspace_map, data_rows, filepath):
    """Configure side car speed dials for devices"""
    print(f"\n{'='*60}")
    print("Configuring Side Car Speed Dials")
    print(f"{'='*60}")
    
    # Read Webex Side Cars sheet
    sidecar_data = read_excel_sheet(filepath, 'Webex Side Cars')
    if not sidecar_data or len(sidecar_data) < 7:
        print("  Skipped: No side car data found")
        return
    
    # Get extensions from rows 4 and 5, column D (index 3)
    target_extensions = []
    for row_idx in [3, 4]:  # Row 4 and 5 (0-indexed)
        if len(sidecar_data) > row_idx and len(sidecar_data[row_idx]) > 3:
            ext = str(sidecar_data[row_idx][3]).strip() if sidecar_data[row_idx][3] else None
            if ext:
                target_extensions.append(ext)
    
    if not target_extensions:
        print("  Skipped: No target extensions found in rows 4-5")
        return
    
    # Build extension to workspace map
    ext_to_workspace = {}
    for row_idx, workspace_id in workspace_map.items():
        row = data_rows[row_idx - 2]
        extension = str(row[4]).strip() if len(row) > 4 else ""
        if extension in target_extensions:
            ext_to_workspace[extension] = workspace_id
    
    # Get device IDs for target workspaces
    device_map = {}
    for extension, workspace_id in ext_to_workspace.items():
        print(f"\n  Fetching device for extension {extension}...")
        devices_result = api.call("GET", f"telephony/config/workspaces/{workspace_id}/devices",
                                 params={"orgId": api.org_id})
        
        if "error" in devices_result:
            print(f"    Warning: Failed to fetch devices - {devices_result['error']}")
            print(f"    Please manually configure side car for extension {extension}")
            continue
        
        devices = devices_result.get('devices', [])
        if devices:
            device_id = devices[0].get('id')
            device_map[extension] = device_id
            print(f"    Found device ID: {device_id}")
        else:
            print(f"    Warning: No devices found")
    
    if not device_map:
        print("\n  Skipped: No devices found for configuration")
        return
    
    # Build speed dial array from rows 7-34, columns C and D
    kem_keys = []
    for row_idx in range(6, 34):  # Rows 7-34 (0-indexed 6-33)
        if len(sidecar_data) <= row_idx:
            break
        
        row = sidecar_data[row_idx]
        label = str(row[2]).strip() if len(row) > 2 and row[2] else None
        value = str(row[3]).strip() if len(row) > 3 and row[3] else None
        
        if label and value:
            kem_keys.append({
                "kemModuleIndex": 1,
                "kemKeyIndex": len(kem_keys) + 1,
                "kemKeyType": "SPEED_DIAL",
                "kemKeyLabel": label,
                "kemKeyValue": value
            })
    
    if not kem_keys:
        print("\n  Skipped: No speed dials found in rows 7-34")
        return
    
    print(f"\n  Found {len(kem_keys)} speed dial entries")
    
    # Configure each device
    for extension, device_id in device_map.items():
        print(f"\n  Configuring device for extension {extension}...")
        
        layout_data = {
            "layoutMode": "CUSTOM",
            "userReorderEnabled": False,
            "lineKeys": [
                {"lineKeyIndex": 1, "lineKeyType": "PRIMARY_LINE"},
                {"lineKeyIndex": 2, "lineKeyType": "OPEN"},
                {"lineKeyIndex": 3, "lineKeyType": "OPEN"},
                {"lineKeyIndex": 4, "lineKeyType": "OPEN"},
                {"lineKeyIndex": 5, "lineKeyType": "OPEN"},
                {"lineKeyIndex": 6, "lineKeyType": "OPEN"}
            ],
            "kemModuleType": "KEM_20_KEYS",
            "kemKeys": kem_keys
        }
        
        result = api.call("PUT", f"telephony/config/devices/{device_id}/layout",
                         data=layout_data, params={"orgId": api.org_id})
        
        if "error" in result:
            print(f"    Warning: Failed to configure - {result['error']}")
            print(f"    Please manually configure side car for extension {extension}")
        else:
            print(f"    Success: Side car speed dials configured")

def process_bulk_import(api, location_data, filepath):
    """Process bulk import of workspaces from Excel file"""
    # Read Webex Users sheet
    users_data = read_excel_sheet(filepath, 'Webex Users')
    if not users_data or len(users_data) < 2:
        print("Error: Could not read data")
        return
    
    headers = users_data[0]
    data_rows = users_data[1:]
    
    # Build preview table
    print(f"\n{'='*80}")
    print("Import Preview")
    print(f"{'='*80}")
    
    users_count = 0
    workspaces_count = 0
    preview_items = []
    
    for row_idx, row in enumerate(data_rows, start=2):
        user_type = str(row[9]).strip().lower() if len(row) > 9 else ""
        display_name = str(row[12]).strip() if len(row) > 12 else "Unknown"  # Column M
        extension = str(row[4]).strip() if len(row) > 4 else ""
        phone_number = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        device_model = str(row[10]).strip() if len(row) > 10 else ""
        
        if user_type == 'user':
            users_count += 1
            preview_items.append({
                'row': row_idx,
                'type': 'User',
                'name': display_name,
                'ext': extension,
                'phone': phone_number,
                'device': device_model
            })
        else:
            workspaces_count += 1
            preview_items.append({
                'row': row_idx,
                'type': 'Workspace',
                'name': display_name,
                'ext': extension,
                'phone': phone_number,
                'device': device_model
            })
    
    # Display table
    print(f"{'Row':<5} {'Type':<10} {'Name':<25} {'Ext':<8} {'Phone':<12} {'Device':<20}")
    print(f"{'-'*80}")
    for item in preview_items:
        print(f"{item['row']:<5} {item['type']:<10} {item['name']:<25} {item['ext']:<8} {item['phone']:<12} {item['device']:<20}")
    
    print(f"\n{'='*80}")
    print(f"Total: {len(preview_items)} items ({users_count} users, {workspaces_count} workspaces)")
    print(f"Note: Users will be skipped (not yet implemented)")
    print(f"{'='*80}")
    
    # Confirm
    confirm = input("\nProceed with import? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Import cancelled.")
        return
    
    # Process each row
    print(f"\n{'='*60}")
    print("Starting Bulk Import Process")
    print(f"{'='*60}")
    
    results = {'users': 0, 'workspaces_created': 0, 'workspaces_failed': 0, 'errors': []}
    workspace_map = {}
    
    for row_idx, row in enumerate(data_rows, start=2):
        user_type = str(row[9]).strip().lower() if len(row) > 9 else ""
        display_name = str(row[12]).strip() if len(row) > 12 else "Unknown"  # Column M
        
        if user_type == 'user':
            results['users'] += 1
            print(f"Row {row_idx}: Skipping user '{display_name}' (user provisioning not yet implemented)")
            continue
        
        # Process non-user (workspace)
        print(f"\nRow {row_idx}: Creating workspace '{display_name}'...")
        workspace_id, error = create_workspace_from_row(api, location_data, row, headers)
        
        if workspace_id:
            results['workspaces_created'] += 1
            workspace_map[row_idx] = workspace_id
            if error:
                print(f"  Warning: {error}")
                results['errors'].append(f"Row {row_idx}: {error}")
            else:
                print(f"  Success: Workspace created (ID: {workspace_id})")
        else:
            results['workspaces_failed'] += 1
            print(f"  Failed: {error}")
            results['errors'].append(f"Row {row_idx}: {error}")
    
    # Configure call forwarding
    if workspace_map:
        print(f"\n{'='*60}")
        print("Configuring Call Forwarding & Business Continuity")
        print(f"{'='*60}")
        
        for row_idx, workspace_id in workspace_map.items():
            row = data_rows[row_idx - 2]
            display_name = str(row[12]).strip() if len(row) > 12 else "Unknown"
            
            print(f"\nRow {row_idx}: Configuring '{display_name}'...")
            error = configure_call_forwarding(api, workspace_id, row)
            
            if error:
                print(f"  Warning: {error}")
                results['errors'].append(f"Row {row_idx}: Call forwarding failed - {error}")
            else:
                print(f"  Success: Call forwarding configured")
    
    # Configure outgoing permissions
    if workspace_map:
        print(f"\n{'='*60}")
        print("Configuring Outgoing Calling Permissions")
        print(f"{'='*60}")
        
        for row_idx, workspace_id in workspace_map.items():
            row = data_rows[row_idx - 2]
            display_name = str(row[12]).strip() if len(row) > 12 else "Unknown"
            
            print(f"\nRow {row_idx}: Checking '{display_name}'...")
            error, was_configured = configure_outgoing_permission(api, workspace_id, row)
            
            if was_configured:
                if error:
                    print(f"  Warning: {error}")
                    results['errors'].append(f"Row {row_idx}: Outgoing permission failed - {error}")
                else:
                    print(f"  Success: Custom outgoing permissions configured")
            else:
                print(f"  Skipped: No custom permissions required")
    
    # Configure side car speed dials
    if workspace_map:
        print(f"\n{'='*60}")
        proceed = input("\nProceed with side car speed dial configuration? (Y/n): ").strip().lower()
        if proceed in ['', 'y', 'yes']:
            configure_side_car_speed_dials(api, workspace_map, data_rows, filepath)
        else:
            print("\nSide car configuration skipped. Workflow complete.")
            return
    
    # Print summary
    print(f"\n{'='*60}")
    print("Bulk Import Summary")
    print(f"{'='*60}")
    print(f"Users skipped: {results['users']}")
    print(f"Workspaces created: {results['workspaces_created']}")
    print(f"Workspaces failed: {results['workspaces_failed']}")
    
    if results['errors']:
        print(f"\nErrors/Warnings:")
        for error in results['errors']:
            print(f"  - {error}")
    
    if results['users'] > 0:
        print(f"\nNote: User provisioning will be implemented in a future update.")
    
    print(f"{'='*60}")

def aso_bulk_import_tool(api):
    """Main function for ASO Bulk Import Tool"""
    print("\n--- ASO Bulk Import Tool ---")
    
    # Check if bulk folder exists
    if not os.path.exists('bulk'):
        print("Status: FAILED - 'bulk' folder not found")
        print("Creating 'bulk' folder...")
        os.makedirs('bulk')
        print("Please place your 'aso_import' Excel file in the 'bulk' folder and try again.")
        return
    
    # Find aso_import file
    print("\nSearching for 'aso_import' Excel file in bulk folder...")
    filepath = find_aso_import_file()
    
    if not filepath:
        print("Status: FAILED - No file found with prefix 'aso_import' (.xlsx or .xls)")
        print("\nPlease ensure your Excel file:")
        print("  1. Has a filename starting with 'aso_import'")
        print("  2. Is in Excel format (.xlsx or .xls)")
        print("  3. Is located in the 'bulk' folder")
        return
    
    print(f"Status: PASS - Found file: {filepath}")
    
    # Validate Excel file
    is_valid, additional_tabs = validate_excel_file(filepath)
    
    if not is_valid:
        print("\nValidation failed. Please fix the issues and try again.")
        return
    
    # Store additional tabs in memory
    if additional_tabs:
        print(f"\nAdditional tabs detected and cached:")
        for i, tab in enumerate(additional_tabs, 1):
            print(f"  {i}. {tab}")
    
    # Validate location
    location = validate_location(api, filepath)
    
    if not location:
        print("\nValidation failed. Returning to previous menu.")
        return
    
    # Validate Webex Users data
    if not validate_webex_users_data(filepath):
        print("\nValidation failed. Returning to previous menu.")
        return
    
    # Validate available phone numbers
    if not validate_available_numbers(api, location, filepath):
        print("\nValidation failed. Returning to previous menu.")
        return
    
    print("\nValidation complete. Ready for next steps.")
    
    # Process bulk import
    process_bulk_import(api, location, filepath)
